"""Inspect the schedule sheet in Excel to see actual data."""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

goals_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "цели_2026_трекер.xlsx")
if not os.path.exists(goals_file):
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    goals_file = os.path.join(desktop, "цели_2026_трекер.xlsx")

print(f"Opening: {goals_file}")
if not os.path.exists(goals_file):
    print("FILE NOT FOUND!")
    sys.exit(1)

import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook(goals_file, data_only=True)
print(f"Sheets: {wb.sheetnames}")

ws = wb["Расписание дня"]
print(f"\n=== Расписание дня (rows 6-70, first 20 rows) ===")
print(f"Sheet dimensions: {ws.dimensions}")
count = 0
for i, row in enumerate(ws.iter_rows(min_row=6, max_row=70, values_only=True), start=6):
    vals = []
    for j, v in enumerate(row[:5]):
        if v is not None:
            vals.append(f"  col{j}={repr(v)[:40]}  type={type(v).__name__}")
    if vals:
        print(f"Row {i}: {''.join(vals)}")
    else:
        print(f"Row {i}: [empty]")
    count += 1
    if count >= 25:
        break

wb.close()
print(f"\nTotal rows checked: {count}")
