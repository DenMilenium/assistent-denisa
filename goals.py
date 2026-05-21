"""
Goals module — reads цели_2026_трекер.xlsx and syncs data
"""

import os
import logging
import openpyxl
from datetime import datetime

logger = logging.getLogger(__name__)

GOALS_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "цели_2026_трекер.xlsx")
# Fallback to desktop
if not os.path.exists(GOALS_FILE):
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    GOALS_FILE = os.path.join(desktop, "цели_2026_трекер.xlsx")


def _parse_progress(val):
    """Parse progress value to float, handling strings and None."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace("%", "").replace(",", ".").strip())
    except (ValueError, TypeError):
        return 0


def load_all_goals():
    """Load goals data from Excel file."""
    if not os.path.exists(GOALS_FILE):
        return {"goals": [], "subtasks": [], "schedule": [], "monthly": [], "billionaires": [], "total_progress": 0}

    try:
        wb = openpyxl.load_workbook(GOALS_FILE, data_only=True)
    except Exception as e:
        logger.warning(f"Failed to load Excel file: {e}")
        return {"goals": [], "subtasks": [], "schedule": [], "monthly": [], "billionaires": [], "total_progress": 0}

    data = {}
    sheet_names = wb.sheetnames

    # General goals
    try:
        ws = wb["Генеральные цели"]
    except KeyError:
        logger.warning("Excel: sheet 'Генеральные цели' not found")
        ws = None
    goals = []
    if ws:
        for row in ws.iter_rows(min_row=7, max_row=12, values_only=True):
            if row and len(row) > 1 and row[1]:
                goals.append({
                    "id": row[0] if row[0] else 0,
                    "name": str(row[1]) if row[1] else "",
                    "deadline": str(row[2]) if len(row) > 2 and row[2] else "",
                    "status": str(row[3]) if len(row) > 3 and row[3] else "",
                    "progress": _parse_progress(row[4]) if len(row) > 4 else 0,
                    "description": str(row[5]) if len(row) > 5 and row[5] else "",
                })
    data["goals"] = goals

    # Sub-tasks
    try:
        ws = wb["Подзадачи"]
    except KeyError:
        logger.warning("Excel: sheet 'Подзадачи' not found")
        ws = None
    subtasks = []
    if ws:
        for row in ws.iter_rows(min_row=7, max_row=42, values_only=True):
            if row and len(row) > 1 and row[1]:
                subtasks.append({
                    "goal_id": row[0] if row[0] else 0,
                    "name": str(row[1]) if row[1] else "",
                    "category": str(row[2]) if len(row) > 2 and row[2] else "",
                    "deadline": str(row[3]) if len(row) > 3 and row[3] else "",
                    "progress": _parse_progress(row[4]) if len(row) > 4 else 0,
                    "status": str(row[5]) if len(row) > 5 and row[5] else "",
                    "description": str(row[6]) if len(row) > 6 and row[6] else "",
                })
    data["subtasks"] = subtasks

    # Daily schedule
    try:
        ws = wb["Расписание дня"]
    except KeyError:
        logger.warning("Excel: sheet 'Расписание дня' not found")
        ws = None
    schedule = []
    if ws:
        for row in ws.iter_rows(min_row=6, max_row=70, values_only=True):
            # Date is in column 2 (row[1]), column 1 (row[0]) is empty or legend
            date_val = None
            if not row:
                continue
            for col_idx in range(min(3, len(row))):
                if row[col_idx] and isinstance(row[col_idx], (datetime, str)):
                    date_val = row[col_idx]
                    break
            if not date_val:
                continue
            # Handle both datetime objects and string dates
            if isinstance(date_val, str):
                try:
                    date_val = datetime.strptime(date_val, "%d.%m.%Y")
                except ValueError:
                    continue
            if isinstance(date_val, datetime):
                schedule.append({
                    "date": date_val,
                    "day": str(row[1]) if len(row) > 1 and row[1] else "",
                    "morning": str(row[2]) if len(row) > 2 and row[2] else "",
                    "sport": str(row[3]) if len(row) > 3 and row[3] else "",
                    "breakfast": str(row[4]) if len(row) > 4 and row[4] else "",
                    "work_morning": str(row[5]) if len(row) > 5 and row[5] else "",
                    "lunch": str(row[6]) if len(row) > 6 and row[6] else "",
                    "work_evening": str(row[7]) if len(row) > 7 and row[7] else "",
                    "dinner": str(row[8]) if len(row) > 8 and row[8] else "",
                    "development": str(row[9]) if len(row) > 9 and row[9] else "",
                    "review": str(row[10]) if len(row) > 10 and row[10] else "",
                    "priority": str(row[11]) if len(row) > 11 and row[11] else "",
                })
    data["schedule"] = schedule

    # Monthly KPIs
    try:
        ws = wb["По месяцам"]
    except KeyError:
        logger.warning("Excel: sheet 'По месяцам' not found")
        ws = None
    monthly = []
    if ws:
        for row in ws.iter_rows(min_row=7, max_row=15, values_only=True):
            if row and row[0]:
                monthly.append({
                    "month": str(row[0]) if row[0] else "",
                    "books_plan": row[1] if len(row) > 1 else None,
                    "books_fact": row[2] if len(row) > 2 else None,
                    "neural": str(row[3]) if len(row) > 3 and row[3] else "",
                    "weight": row[4] if len(row) > 4 else None,
                    "income": row[5] if len(row) > 5 and row[5] else 0,
                    "english": str(row[6]) if len(row) > 6 and row[6] else "",
                    "contacts": row[7] if len(row) > 7 else 0,
                    "contracts": str(row[8]) if len(row) > 8 and row[8] else "",
                })
    data["monthly"] = monthly

    # Billionaires
    try:
        ws = wb["Миллиардеры"]
    except KeyError:
        logger.warning("Excel: sheet 'Миллиардеры' not found")
        ws = None
    billionaires = []
    if ws:
        for row in ws.iter_rows(min_row=7, max_row=16, values_only=True):
            if row and len(row) > 1 and row[1]:
                billionaires.append({
                    "num": row[0],
                    "name": str(row[1]) if row[1] else "",
                    "company": str(row[2]) if len(row) > 2 and row[2] else "",
                    "country": str(row[3]) if len(row) > 3 and row[3] else "",
                    "field": str(row[4]) if len(row) > 4 and row[4] else "",
                    "strategy": str(row[5]) if len(row) > 5 and row[5] else "",
                    "status": str(row[6]) if len(row) > 6 and row[6] else "",
                })
    data["billionaires"] = billionaires

    # Dashboard metrics
    total_progress = 0
    goal_count = 0
    for g in goals:
        total_progress += g["progress"]
        goal_count += 1
    data["total_progress"] = round(total_progress / goal_count, 1) if goal_count else 0

    wb.close()
    return data


def get_today_schedule():
    """Get today's schedule from goals file."""
    data = load_all_goals()
    today = datetime.now().date()
    for item in data.get("schedule", []):
        if isinstance(item.get("date"), datetime) and item["date"].date() == today:
            return item
    return None


def get_urgent_tasks():
    """Get tasks that are due within 7 days or overdue."""
    data = load_all_goals()
    today = datetime.now()
    urgent = []

    for task in data.get("subtasks", []):
        status = task.get("status", "").strip().lower()
        if status in ("выполнено", "выполнено"):
            continue
        if task.get("progress", 0) >= 100:
            continue

        deadline_str = task.get("deadline", "").strip()
        if not deadline_str:
            continue

        # Parse various date formats
        deadline = None
        for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d.%m.%Y"]:
            try:
                deadline = datetime.strptime(deadline_str, fmt)
                break
            except ValueError:
                continue

        if deadline is None:
            # Try month-year format
            for fmt in ["%B %Y", "%b %Y", "%m.%Y"]:
                try:
                    deadline = datetime.strptime(deadline_str, fmt)
                    break
                except ValueError:
                    continue

        if deadline:
            days_left = (deadline - today).days
            if -30 <= days_left <= 7:
                task["days_left"] = days_left
                urgent.append(task)

    return sorted(urgent, key=lambda x: x.get("days_left", 999))
